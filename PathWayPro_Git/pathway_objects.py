import os
import pandas as pd
from openpyxl import load_workbook
import shutil
import datetime

working_dir = os.getcwd() + "\\"
csv_file_path = 'master_rep.csv'
csv_data = pd.read_csv(csv_file_path)
excel_file_path = 'DDS hits for 2024.xlsx'
sheet_name = 'Summer calibrations'
backup_folder = "DDS BackUps"


class FileProcessor:
    def __init__(self, unprocessed_dir="unprocessed", processed_dir="processed", processed_excel = "processed_excel"):
        self.unprocessed_dir = unprocessed_dir
        self.processed_dir = processed_dir
        self.processed_excel = processed_excel
        self.db_list = []
    
    

    def backup_excel_file(self, original_file=excel_file_path, backup_folder=backup_folder):
        """Creates a daily backup file for the main Excel file used and date-stamps it"""
        date_stamp = datetime.datetime.now().strftime("%m-%d-%Y")
        backup_file = os.path.join(backup_folder, f"DDS hits for 2024_{date_stamp}.xlsx")
        
        # Ensure backup folder exists
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)
        
        # Check if today's backup already exists
        if os.path.exists(backup_file):
            return f"""Backup for today ({date_stamp})already exists: 
{backup_file}

Make Sure You Have Saved New Files to the 
'unprocessed' Folder
****Proceed with Processing Files****"""
        
        # Copy the file
        shutil.copy2(original_file, backup_file)
        
        return f"""Backup created successfully: 
{backup_file}

Make Sure You Have Saved New Files to the 
'unprocessed' Folder
*****Proceed with Processing Files*****"""

    def process(self):
        """Takes .xlsx files from one folder, converts them to .csv, and moves them to another.
        Default directory paths are 'unprocessed' -> 'processed'."""
        
        processed_list = []
        # Check if the unprocessed directory exists
        if not os.path.exists(self.unprocessed_dir):
            return
        
        # Create the processed directory if it does not exist
        if not os.path.exists(self.processed_dir):
            os.makedirs(self.processed_dir)
        
        # Iterate through each file in the unprocessed directory
        for filename in os.listdir(self.unprocessed_dir):
            file_path = os.path.join(self.unprocessed_dir, filename)
            
            # Check if it's a file
            if os.path.isfile(file_path):
                # Read the file into a DataFrame
                try:
                    df = pd.read_excel(file_path)  # Assuming the files are Excel files
                except Exception as e:
                    continue
                
                # Define the new file name with .csv extension
                new_filename = os.path.splitext(filename)[0] + '.csv'
                new_file_path = os.path.join(self.processed_dir, new_filename)
                processed_list.append(new_file_path)
                
                # Save the DataFrame as a .csv file
                try:
                    df.to_csv(new_file_path, index=False)
                    self.db_list.append(new_filename)
                except Exception as e:
                    continue
        return processed_list


    def move_and_delete_excel_files(self):
        # Define the source and destination directories
        source_dir = 'unprocessed'
        dest_dir = 'processed_excel'

        # Ensure the destination directory exists
        if not os.path.exists(dest_dir):
            os.makedirs(dest_dir)

        # Get all files in the source directory
        files = os.listdir(source_dir)

        # Iterate over the files
        for file_name in files:
            # Construct full file path
            source_file = os.path.join(source_dir, file_name)

            # Check if the file is an Excel file
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                # Construct the destination file path
                dest_file = os.path.join(dest_dir, file_name)

                # Move the file
                shutil.move(source_file, dest_file)
                print(f'Moved: {file_name}')

        print('All Excel files have been moved and the originals deleted.')



    
    def merge(self):
        """Combines all the .csv files in the processed folder into a csv file in the working directory"""
        # Initialize an empty DataFrame to store the merged data
        master_df = pd.DataFrame()
        checklist = []

        # Define the subdirectory where processed files are located
        subdirectory = "processed"

        # Get the current working directory
        current_dir = os.getcwd()

        # Construct the path to the processed subdirectory
        processed_dir = os.path.join(current_dir, subdirectory)
        
        # Loop through files in the processed subdirectory
        for file in os.listdir(processed_dir):
            if file.endswith(".csv"):
                # Construct the full file path
                file_path = os.path.join(processed_dir, file)
                # Read each CSV file and append to master_df
                master_df = master_df._append(pd.read_csv(file_path))
                checklist.append(file_path + "\n")
        
        # Save the merged DataFrame to a new CSV file
        master_df.to_csv('master_rep.csv', index=False)
        return checklist

     
    def add_to_excel(self, csv_file_path=csv_file_path, excel_file_path=excel_file_path, sheet_name=sheet_name):
        """csv_file_path = path to data to add to excel sheet
        excel_file_path = path to actual workbook
        sheet_name = string of the name of the sheet within the excel file you want to add to"""
        # Read the CSV File (assuming 'master_rep.csv' exists)
        df_csv = pd.read_csv(csv_file_path)

        # Load the Excel File (create if not exists)
        excel_file = excel_file_path

        try:
            # Load workbook
            wb = load_workbook(excel_file_path)
        except FileNotFoundError:
            # If file does not exist, create new workbook
            wb = Workbook()

        # Append CSV Data to Excel Sheet
        sheet_name = 'Summer calibrations'

        # Check if sheet exists, create if not
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Append data from df_csv to existing sheet
        rows = df_csv.values.tolist()
        for row in rows:
            ws.append(row)
        
        # Delete empty rows
        # Iterate from last row to first row to avoid skipping rows after deletion
        for row in range(ws.max_row, 0, -1):
            if all(cell.value is None for cell in ws[row]):
                ws.delete_rows(row, 1)

        # Save workbook
        wb.save(excel_file)
        saved = f"{csv_file_path} has been added to \nthe {sheet_name} sheet in  \n{excel_file_path}"
        return saved


    def reset_master_rep(self, filename=csv_file_path):
       
        # Delete the file if it exists
        if os.path.exists(filename):
            os.remove(filename)
        
        # Delete processed folder if it exists, quick way to avoid repeating data inputs
        folder_name = 'processed'
    
    # Construct the full folder path
        folder_path = os.path.join(working_dir, folder_name)
    
    # Check if the folder exists
        if os.path.exists(folder_path) and os.path.isdir(folder_path):
            # Delete the folder and its contents
            shutil.rmtree(folder_path)

        # Define the headers
        headers = [
            '__AutoNum', '_Close_and_latch_MCB_Cover', '_Install Lid Vent', '_Open_and_close_ball_valves', '_Remote_Type_or_SP#',
            'Add RV antifreeze to water line', 'Address1', 'Address2', 'AMM', 'Anticipation', 'Attach Clamps', 'Auger Moter Amp',
            'Ball Valve #1 Serial Number', 'Ball Valve #2 Serial Number', 'Barcode', 'Batch', 'BatchDate', 'Blow water out of water line',
            'Cal #1:  _Actual_', 'Cal #1:  _Setpoint_', 'Cal #2:  _Actual_', 'Cal #2:  _Setpoint_', 'Cal #3:  _Actual_', 'Cal #3:  _Setpoint_',
            'Cell Phone', 'CF AFTER Calibration', 'CF After Calibration Don\'t Use', 'CF Upon Arrival', 'Check Lid Gasket', 'City',
            'Clean _ lube vibrator relay', 'Clean _ paint auger motor', 'Clean _ paint pump flanges', 'Clean _ paint pump motor', 'Clean Paint Vib',
            'Comments', 'Company Name', 'Contact', 'Description of problem', 'Empty Product from unit', 'End Time', 'EndAct', 'FarmLoc', 'Fax Number',
            'GND to L1', 'GND to L2', 'GPS Form', 'GPS Unit', 'Inlet Screen', 'Install cap', 'Install Lid Vent', 'Install plastic saran cover on funnel',
            'Install plug', 'Kahler Conversion', 'L1 to L3', 'MM', 'Notes', 'Overall condition of unit upon arrival', 'Paint Auger Motor', 'Paint Pump',
            'Part Replaced Condition _Def_Dmg_ _1_', 'Part Replaced Condition _Def_Dmg_ _2_', 'Part Replaced Condition _Def_Dmg_ _3_',
            'Part Replaced Condition _Def_Dmg_ _4_', 'Part Replaced Condition _Def_Dmg_ _5_', 'Part Replaced Condition _Def_Dmg_ _6_',
            'Part Replaced Condition _Def_Dmg_ _7_', 'Part Replaced Description _1_', 'Part Replaced Description _2_', 'Part Replaced Description _3_',
            'Part Replaced Description _4_', 'Part Replaced Serial # _1_', 'Part Replaced Serial # _2_', 'Part Replaced Serial # _3_', 'Part Replaced Serial # _4_',
            'Parts Replaced', 'Parts Replaced 5', 'Parts Replaced 6', 'Parts Replaced 7', 'Parts Replaced Serial 5', 'Parts Replaced Serial 6',
            'Parts Replaced Serial 7', 'Patent Stick on Unit', 'Pic1GPS', 'Pic2GPS', 'Pic3GPS', 'Pic4GPS', 'Pic5GPS', 'Pic6GPS', 'Power Supply Output',
            'Printer', 'Problem Description', 'Product', 'Pump PSI', 'Pump_Motor Serial Number', 'Remote', 'Replace Power Supply', 'Reset Inventory',
            'Reset Inventory to 0', 'Run Water', 'Running Total 2', 'Running Total 3', 'Running Total 4', 'RunningAfter2', 'RunningAfter3', 'RunningAfter4',
            'Salesman', 'Scale _Platform_Cell_No_', 'Scale Unit_Load Cell', 'Service Date', 'Service Type', 'Single Flight Auger', 'Software Version', 'Source',
            'Speed Control Output', 'Spray lube upper pump seal', 'Start Time', 'StartAct', 'State_Province', 'Technician', 'Telephone', 'Transformer',
            'Unit Accessibility', 'Unit Inventory', 'Unit Number', 'Vibrator Amp', 'Warning Label _MCB_', 'Wash Butyrate Elbow', 'Wash DDS floor', 'Wash DDS product bin',
            'Wash Unit', 'Wildcard', 'Zip Code'
        ]

        # Create an empty DataFrame with the specified headers
        df = pd.DataFrame(columns=headers)
        
        # Save the DataFrame to a CSV file
        df.to_csv(filename, index=False)


